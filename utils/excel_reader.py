import openpyxl
from utils.logger import get_logger

logger = get_logger("ExcelReader")

def read_login_data(
    file_path="test_data/login_data.xlsx",
    sheet_name="LoginData"
):
    logger.info(f"Reading Excel file: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    data = []
    headers = [cell.value for cell in ws[1]]  # Row 1 = headers

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # skip empty rows
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
            logger.info(f"Loaded row: {row_dict}")

    logger.info(f"Total rows loaded: {len(data)}")
    return data